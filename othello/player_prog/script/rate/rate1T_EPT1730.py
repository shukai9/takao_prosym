# python 3
# -*- coding: utf-8 -*-
# シードを10から100まで+10ずつ
# レート10種類をテキストファイルに出力し, その平均をエクセルに入れる
import re
import openpyxl as px
import numpy as np

board_data = [130, 452, 789, 680, 891, 544, 954, 995, 361, 673, 52, 399, 417, 631, 648, 714, 390, 615, 566, 401, 731, 596, 162, 745, 827, 708, 514, 113, 541, 675, 348, 686, 172, 259, 216, 907, 804, 826, 246, 261, 264, 886, 569, 897, 837, 391, 872, 156, 379, 252, 749, 195, 257, 238, 654, 658, 126, 284, 151, 314, 611, 770, 328, 823, 6, 666, 967, 485, 54, 155, 169, 208, 733, 817, 546, 486, 774, 682, 174, 958, 112, 769, 148, 664, 972, 5, 454, 455, 80, 429, 16, 46, 198, 549, 313, 423, 735, 281, 178, 592]
player = ["random100", "hyouka-0-100", "MC-100", "MCT", "EPT_MCT1730", "EPT_MCT8500", "AlphaBeta1", "AlphaBeta2", "AlphaBeta4"]
player_len = len(player)
name = "EPT1730"

wb = px.Workbook()
ws = wb.active
for i in range(player_len):
    ws.cell(row=1, column=i+2, value=player[i])
for i in range(100):
    ws.cell(row=i+2, column=1, value=board_data[i])
for i in range(100):
    rate = [[0 for x in range(9)] for y in range(10)]
    file = open("../result/board1rateT_{}/board{}.txt".format(name, board_data[i]), "r")
    string = file.readlines()
    for count in range(10):
        for cpu_num in range(10):
            if(cpu_num==0):
                continue
            data = string[(count*10)+cpu_num].split(":")[1]
            data = data.replace(" ", "")
            data = data.replace("\n", "")
            data = data.replace("\r", "")
            print("i : {}, count : {}, cpu_num : {}".format(i, count, cpu_num))
            rate[count][cpu_num-1] = float(data)
    ave = np.mean(rate, axis=0)
    for x in range(player_len):
        ws.cell(row=i+2, column=x+2, value=ave[x])
wb.save('../result/board1rateT_{}/excel/{}.xlsx'.format(name, name))
file.close()
