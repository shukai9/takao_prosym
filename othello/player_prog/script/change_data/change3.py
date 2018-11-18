# python 3
# -*- coding: utf-8 -*-
import re
import openpyxl as px

player = ["random100", "hyouka-0-100", "MC-100", "MCT", "EPT_MCT1730", "EPT_MCT8500", "AlphaBeta1", "AlphaBeta2", "AlphaBeta4"]
player_len = len(player)
board_data = [624, 811, 575, 961, 883, 179, 260, 500, 688, 824, 165, 509, 555, 520, 122, 18, 109, 285, 413, 277, 981, 48, 402, 414, 642, 944, 767, 783, 488, 209, 467, 639, 764, 947, 86, 427, 501, 743, 918, 986, 24, 371, 623, 668, 748, 807, 217, 638, 495, 254, 604, 618, 57, 202, 2, 190, 881, 26, 158, 607, 819, 715, 916, 517, 788, 840, 177, 171, 591, 818, 72, 558, 822, 874, 424, 489, 1000, 862, 554, 97, 780, 908, 49, 563, 943, 168, 337, 539, 61, 468, 550, 829, 327, 752, 800, 849, 890, 349, 633, 828]


for boardnum in range(100):
    file = open("result/board3copy/board{}.txt".format(board_data[boardnum]), "r")
    string = file.readlines()
    wb = px.Workbook()
    ws = wb.active
    ws['A1'].value = 'Board{}'.format(board_data[boardnum])

    for i in range(player_len):
        ws.cell(row=1, column=i+2, value=player[i])

        count = 0
        for j in range(player_len):
            ws.cell(row=j+2, column=1, value=player[j])
            for i in range(player_len):
                data = string[count].split(":")[1]
                data = data.replace(" ", "")
                data = data.replace("\n", "")
                data = data.replace("\r", "")
                ws.cell(row=j+2, column=i+2, value=data)
                count+=1
    wb.save('result/board3/excel/board{}.xlsx'.format(board_data[boardnum]))
    file.close()
