# python 3
# -*- coding: utf-8 -*-
import re
import openpyxl as px

player = ["random100", "hyouka-0-100", "MC-100", "MCT", "EPT_MCT1730", "EPT_MCT8500", "AlphaBeta1", "AlphaBeta2", "AlphaBeta4"]
player_len = len(player)
board_data = [130, 452, 789, 680, 891, 544, 954, 995, 361, 673, 52, 399, 417, 631, 648, 714, 390, 615, 566, 401, 731, 596, 162, 745, 827, 708, 514, 113, 541, 675, 348, 686, 172, 259, 216, 907, 804, 826, 246, 261, 264, 886, 569, 897, 837, 391, 872, 156, 379, 252, 749, 195, 257, 238, 654, 658, 126, 284, 151, 314, 611, 770, 328, 823, 6, 666, 967, 485, 54, 155, 169, 208, 733, 817, 546, 486, 774, 682, 174, 958, 112, 769, 148, 664, 972, 5, 454, 455, 80, 429, 16, 46, 198, 549, 313, 423, 735, 281, 178, 592]


for boardnum in range(100):
    file = open("result/board1copy/board{}.txt".format(board_data[boardnum]), "r")
    string = file.readlines()
    wb = px.Workbook()
    ws = wb.active
    ws['A1'].value = 'Board{}'.format(board_data[boardnum])

    for i in range(player_len):
        ws.cell(row=1, column=i+2, value=player[i])

        count = 0
        for j in range(player_len):
            ws.cell(row=j+2, column=1, value=player[j])
            for i in range(player_len):
                data = string[count].split(":")[1]
                data = data.replace(" ", "")
                data = data.replace("\n", "")
                data = data.replace("\r", "")
                ws.cell(row=j+2, column=i+2, value=data)
                count+=1
    wb.save('result/board1/excel/board{}.xlsx'.format(board_data[boardnum]))
    file.close()
