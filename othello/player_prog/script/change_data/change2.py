# python 3
# -*- coding: utf-8 -*-
import re
import openpyxl as px

player = ["random100", "hyouka-0-100", "MC-100", "MCT", "EPT_MCT1730", "EPT_MCT8500", "AlphaBeta1", "AlphaBeta2", "AlphaBeta4"]
player_len = len(player)
board_data = [641, 795, 433, 411, 184, 530, 334, 798, 722, 204, 505, 587, 215, 82, 453, 856, 201, 598, 676, 74, 616, 719, 980, 600, 123, 583, 432, 994, 333, 585, 464, 4, 53, 738, 812, 70, 69, 442, 537, 644, 963, 683, 388, 133, 354, 861, 152, 335, 465, 904, 953, 64, 230, 832, 834, 568, 778, 135, 450, 289, 504, 239, 750, 866, 106, 186, 83, 574, 605, 677, 900, 225, 229, 562, 408, 496, 237, 311, 702, 761, 56, 241, 243, 502, 988, 197, 111, 33, 393, 398, 338, 622, 326, 894, 370, 940, 991, 493, 137, 430]


for boardnum in range(100):
    file = open("result/board2copy/board{}.txt".format(board_data[boardnum]), "r")
    string = file.readlines()
    wb = px.Workbook()
    ws = wb.active
    ws['A1'].value = 'Board{}'.format(board_data[boardnum])

    for i in range(player_len):
        ws.cell(row=1, column=i+2, value=player[i])

        count = 0
        for j in range(player_len):
            ws.cell(row=j+2, column=1, value=player[j])
            for i in range(player_len):
                data = string[count].split(":")[1]
                data = data.replace(" ", "")
                data = data.replace("\n", "")
                data = data.replace("\r", "")
                ws.cell(row=j+2, column=i+2, value=data)
                count+=1
    wb.save('result/board2/excel/board{}.xlsx'.format(board_data[boardnum]))
    file.close()
